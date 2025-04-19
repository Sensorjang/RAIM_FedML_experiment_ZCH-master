from raim_torch_hierarchicalfl_step_by_step_exp import raim
from raim_rs_torch_hierarchicalfl_step_by_step_exp import raim_rs
import openpyxl
import time
import matplotlib.pyplot as plt

excel_save_path = "D:\\"

# 绘图数据
# figure1: 随着ednum的增加(40,50,60,70,80,90,100), su的值
raimrs_su_list_ib_ednum = [] # 随着ednum的增加，raim_su的值
raim_su_list_ib_ednum = [] # 随着ednum的增加，raimrs_su的值

# figure2: 随着esnum的增加(5,10,15,20,25,30,35), su的值
raim_su_list_ib_esnum = [] # 随着esnum的增加，raim_su的值
raimrs_su_list_ib_esnum = [] # 随着esnum的增加，raimrs_su的值

#  figure3: 随着ednum的增加(40,50,60,70,80,90,100), cu的值
raimrs_cu_list_ib_ednum = [] # 随着ednum的增加，raim_su的值
raim_cu_list_ib_ednum = [] # 随着ednum的增加，raimrs_su的值

# figure4: 随着esnum的增加(5,10,15,20,25,30,35), cu的值
raimrs_cu_list_ib_esnum = [] # 随着esnum的增加，raim_su的值
raim_cu_list_ib_esnum = [] # 随着esnum的增加，raimrs_su的值

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 创建工作表
fig1 = wb.create_sheet('Figure 1')
fig1.append(['ednum', '[RAIM] social utility', '[RAIM-RS] social utility'])
fig2 = wb.create_sheet('Figure 2')
fig2.append(['esnum', '[RAIM] social utility', '[RAIM-RS] social utility'])
fig3 = wb.create_sheet('Figure 3')
fig3.append(['ednum', '[RAIM] cloudserver utility', '[RAIM-RS] cloudserver utility'])
fig4 = wb.create_sheet('Figure 4')
fig4.append(['esnum', '[RAIM] cloudserver utility', '[RAIM-RS] cloudserver utility'])

if __name__ == '__main__':
    esnum = 15
    for ednum in range(40, 101, 10): # ednum = 40,50,60,70,80,90,100
        raim_su, raim_cu = raim(True, esnum, ednum, 0.3)
        print("[RAIM] sdnum:{} esnum:{} raim_su{} raim_cu{}".format(ednum, esnum, raim_su, raim_cu))
        raimrs_su, raimrs_cu = raim_rs(True, esnum, ednum, 0.3)
        print("[RAIM-RS] sdnum:{} esnum:{} raimrs_su{} raimrs_cu{}".format(ednum, esnum, raimrs_su, raimrs_cu))
        
        # 记录数据
        raim_su_list_ib_ednum.append(raim_su)
        raimrs_su_list_ib_ednum.append(raimrs_su)
        raim_cu_list_ib_ednum.append(raim_cu)
        raimrs_cu_list_ib_ednum.append(raimrs_cu)

    ednum = 70
    for esnum in range(5, 36, 5): # esnum = 5,10,15,20,25,30,35
        raim_su, raim_cu = raim(True, esnum, ednum, 0.3)
        print("[RAIM] sdnum:{} esnum:{} raim_su{} raim_cu{}".format(ednum, esnum, raim_su, raim_cu))
        raimrs_su, raimrs_cu = raim_rs(True, esnum, ednum, 0.3)
        print("[RAIM-RS] sdnum:{} esnum:{} raimrs_su{} raimrs_cu{}".format(ednum, esnum, raimrs_su, raimrs_cu))

        # 记录数据
        raim_su_list_ib_esnum.append(raim_su)
        raimrs_su_list_ib_esnum.append(raimrs_su)
        raim_cu_list_ib_esnum.append(raim_cu)
        raimrs_cu_list_ib_esnum.append(raimrs_cu)

    # 输出数据到excel表格的四个工作表
    for i in range(len(raim_su_list_ib_ednum)):
        fig1.append([i * 10 + 40, raim_su_list_ib_ednum[i], raimrs_su_list_ib_ednum[i]])
        fig2.append([i * 5 + 5, raim_su_list_ib_esnum[i], raimrs_su_list_ib_esnum[i]])
    
    for i in range(len(raim_cu_list_ib_ednum)):
        fig3.append([i * 10 + 40, raim_cu_list_ib_ednum[i], raimrs_cu_list_ib_ednum[i]])
        fig4.append([i * 5 + 5, raim_cu_list_ib_esnum[i], raimrs_cu_list_ib_esnum[i]])
    
    wb.save(excel_save_path + 'exp2'+ '_' + str(int(time.time())) + '.xlsx')

    # 创建一个包含四个子图的图形 绘制柱状图表并显示在屏幕上
    ednums = list(range(40, 101, 10))  # ednum的值
    esnums = list(range(5, 36, 5))    # esnum的值

    # 创建一个包含四个子图的图形
    fig, axs = plt.subplots(2, 2, figsize=(15, 10))  # 2行2列的子图布局

    # 第一个子图：随着ednum的增加，su的值
    axs[0, 0].plot(ednums, raim_su_list_ib_ednum, marker='o', label='RAIM SU', linestyle='-', linewidth=2, markersize=6)
    axs[0, 0].plot(ednums, raimrs_su_list_ib_ednum, marker='s', label='RAIM-RS SU', linestyle='-', linewidth=2, markersize=6)
    axs[0, 0].set_xlabel('ednum')
    axs[0, 0].set_ylabel('Social Utility (SU)')
    axs[0, 0].set_title('Social Utility vs. ednum')
    axs[0, 0].legend()
    axs[0, 0].grid(True)

    # 第二个子图：随着esnum的增加，su的值
    axs[0, 1].plot(esnums, raim_su_list_ib_esnum, marker='o', label='RAIM SU', linestyle='-', linewidth=2, markersize=6)
    axs[0, 1].plot(esnums, raimrs_su_list_ib_esnum, marker='s', label='RAIM-RS SU', linestyle='-', linewidth=2, markersize=6)
    axs[0, 1].set_xlabel('esnum')
    axs[0, 1].set_ylabel('Social Utility (SU)')
    axs[0, 1].set_title('Social Utility vs. esnum')
    axs[0, 1].legend()
    axs[0, 1].grid(True)

    # 第三个子图：随着ednum的增加，cu的值
    axs[1, 0].plot(ednums, raim_cu_list_ib_ednum, marker='o', label='RAIM CU', linestyle='-', linewidth=2, markersize=6)
    axs[1, 0].plot(ednums, raimrs_cu_list_ib_ednum, marker='s', label='RAIM-RS CU', linestyle='-', linewidth=2, markersize=6)
    axs[1, 0].set_xlabel('ednum')
    axs[1, 0].set_ylabel('Cloudserver Utility (CU)')
    axs[1, 0].set_title('Cloudserver Utility vs. ednum')
    axs[1, 0].legend()
    axs[1, 0].grid(True)
    axs[1, 0].set_ylim(10, 11.5)

    # 第四个子图：随着esnum的增加，cu的值
    axs[1, 1].plot(esnums, raim_cu_list_ib_esnum, marker='o', label='RAIM CU', linestyle='-', linewidth=2, markersize=6)
    axs[1, 1].plot(esnums, raimrs_cu_list_ib_esnum, marker='s', label='RAIM-RS CU', linestyle='-', linewidth=2, markersize=6)
    axs[1, 1].set_xlabel('esnum')
    axs[1, 1].set_ylabel('Cloudserver Utility (CU)')
    axs[1, 1].set_title('Cloudserver Utility vs. esnum')
    axs[1, 1].legend()
    axs[1, 1].grid(True)
    axs[1, 1].set_ylim(10, 11.5)

    # 调整子图之间的间距
    plt.tight_layout()
    plt.savefig(excel_save_path + 'exp2'+ '_' + str(int(time.time())) + '.png', dpi=300)
    # 显示图形
    plt.show()
    time.sleep(1)